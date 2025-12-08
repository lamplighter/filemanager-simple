#!/usr/bin/env python3
"""
Content Analysis Script for File Organization

Extracts structured metadata from files to enable content-based (not filename-based)
organization decisions.

Usage:
    python scripts/analyze_content.py <file_path> [--max-lines 50]
    ./scripts/analyze_content.sh <file_path>

Output: JSON with extracted metadata including:
    - file_type: detected file type
    - content_summary: human-readable summary of contents
    - detected_entities: list of known entities found (TD, Rogers, etc.)
    - detected_dates: list of dates found in content
    - suggested_category: category hint based on content
    - confidence_boost: additional confidence points from content analysis
"""

import argparse
import csv
import json
import os
import re
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional

# Optional imports with graceful fallback
try:
    import PyPDF2
    HAS_PDF = True
except ImportError:
    HAS_PDF = False

try:
    import openpyxl
    HAS_XLSX = True
except ImportError:
    HAS_XLSX = False

try:
    import docx
    HAS_DOCX = True
except ImportError:
    HAS_DOCX = False

try:
    import yaml
    HAS_YAML = True
except ImportError:
    HAS_YAML = False


def load_entity_config() -> Dict[str, Any]:
    """Load entity detection patterns from config.yaml."""
    script_dir = Path(__file__).parent
    config_path = script_dir.parent / "config.yaml"

    if not config_path.exists():
        return {}

    if HAS_YAML:
        with open(config_path, 'r') as f:
            config = yaml.safe_load(f)
            return config.get('entities', {})
    else:
        # Simple fallback: hardcoded patterns if yaml not available
        return {
            "financial": {
                "banks": [
                    {"name": "TD", "patterns": ["TD Canada Trust", "TD Bank", "Toronto-Dominion"]},
                    {"name": "RBC", "patterns": ["Royal Bank", "RBC"]},
                ]
            },
            "utilities": {
                "telecom": [
                    {"name": "Rogers", "patterns": ["Rogers Communications", "Rogers Wireless"]},
                    {"name": "Bell", "patterns": ["Bell Canada", "Bell Mobility"]},
                ]
            },
            "work": [
                {"name": "Uken", "patterns": ["Uken Games", "Uken Studios", "UKEN"]}
            ]
        }


def detect_entities(text: str, entities_config: Dict[str, Any]) -> List[str]:
    """Find known entities in text content."""
    found_entities = []
    text_upper = text.upper()

    def check_patterns(items: List[Dict], name_key: str = "name"):
        for item in items:
            name = item.get(name_key) or item.get("short_name", "")
            patterns = item.get("patterns", [])
            for pattern in patterns:
                if pattern.upper() in text_upper:
                    if name and name not in found_entities:
                        found_entities.append(name)
                    break

    # Check each category
    for category, subcategories in entities_config.items():
        if isinstance(subcategories, list):
            # Direct list of entities (e.g., insurance, work, properties)
            check_patterns(subcategories)
        elif isinstance(subcategories, dict):
            # Nested subcategories (e.g., financial.banks)
            for subcat_name, items in subcategories.items():
                if isinstance(items, list):
                    check_patterns(items)

    return found_entities


def detect_dates(text: str) -> List[str]:
    """Extract dates from text in various formats."""
    dates = []

    # Common date patterns
    patterns = [
        # YYYY-MM-DD
        r'\b(\d{4})-(\d{2})-(\d{2})\b',
        # MM/DD/YYYY or DD/MM/YYYY
        r'\b(\d{1,2})/(\d{1,2})/(\d{4})\b',
        # Month DD, YYYY
        r'\b(January|February|March|April|May|June|July|August|September|October|November|December)\s+(\d{1,2}),?\s+(\d{4})\b',
        # DD Month YYYY
        r'\b(\d{1,2})\s+(January|February|March|April|May|June|July|August|September|October|November|December)\s+(\d{4})\b',
        # Month YYYY
        r'\b(January|February|March|April|May|June|July|August|September|October|November|December)\s+(\d{4})\b',
    ]

    for pattern in patterns:
        matches = re.findall(pattern, text, re.IGNORECASE)
        for match in matches:
            if isinstance(match, tuple):
                date_str = " ".join(str(m) for m in match)
            else:
                date_str = match
            if date_str and date_str not in dates:
                dates.append(date_str)

    # Limit to first 5 dates to avoid noise
    return dates[:5]


def analyze_csv(path: Path, max_lines: int = 50) -> Dict[str, Any]:
    """Analyze CSV file structure and content."""
    result = {
        "headers": [],
        "row_count": 0,
        "sample_rows": [],
        "delimiter": ",",
        "full_text": ""
    }

    try:
        # First, detect delimiter
        with open(path, 'r', encoding='utf-8', errors='replace') as f:
            sample = f.read(4096)

        # Try to detect delimiter
        for delim in [',', '\t', ';', '|']:
            if delim in sample:
                result["delimiter"] = delim
                break

        # Now read the CSV properly
        with open(path, 'r', encoding='utf-8', errors='replace') as f:
            # Store full text for entity detection
            f.seek(0)
            lines = f.readlines()[:max_lines]
            result["full_text"] = "".join(lines)

            f.seek(0)
            reader = csv.reader(f, delimiter=result["delimiter"])

            rows = list(reader)
            if rows:
                result["headers"] = rows[0]
                result["row_count"] = len(rows) - 1  # Exclude header
                result["sample_rows"] = rows[1:6]  # First 5 data rows

    except Exception as e:
        result["error"] = str(e)

    return result


def analyze_pdf(path: Path, max_pages: int = 3) -> Dict[str, Any]:
    """Extract text and metadata from PDF."""
    result = {
        "page_count": 0,
        "text_sample": "",
        "full_text": ""
    }

    if not HAS_PDF:
        result["error"] = "PyPDF2 not installed"
        return result

    try:
        with open(path, 'rb') as f:
            reader = PyPDF2.PdfReader(f)
            result["page_count"] = len(reader.pages)

            text_parts = []
            for page_num in range(min(max_pages, len(reader.pages))):
                page_text = reader.pages[page_num].extract_text()
                if page_text:
                    text_parts.append(page_text)

            result["full_text"] = "\n".join(text_parts)
            # Create a shorter sample for display
            result["text_sample"] = result["full_text"][:1000] if result["full_text"] else ""

    except Exception as e:
        result["error"] = str(e)

    return result


def analyze_text(path: Path, max_lines: int = 50) -> Dict[str, Any]:
    """Analyze plain text files."""
    result = {
        "line_count": 0,
        "text_sample": "",
        "full_text": ""
    }

    try:
        with open(path, 'r', encoding='utf-8', errors='replace') as f:
            lines = f.readlines()
            result["line_count"] = len(lines)
            result["full_text"] = "".join(lines[:max_lines])
            result["text_sample"] = result["full_text"][:1000]

    except Exception as e:
        result["error"] = str(e)

    return result


def analyze_xlsx(path: Path) -> Dict[str, Any]:
    """Analyze Excel spreadsheet."""
    result = {
        "sheet_names": [],
        "sheets": {},
        "full_text": ""
    }

    if not HAS_XLSX:
        result["error"] = "openpyxl not installed"
        return result

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        result["sheet_names"] = wb.sheetnames

        text_parts = []
        for sheet_name in wb.sheetnames[:3]:  # First 3 sheets
            sheet = wb[sheet_name]
            sheet_data = {
                "headers": [],
                "row_count": 0,
                "sample_rows": []
            }

            rows = []
            for i, row in enumerate(sheet.iter_rows(max_row=10, values_only=True)):
                row_values = [str(cell) if cell is not None else "" for cell in row]
                if i == 0:
                    sheet_data["headers"] = row_values
                else:
                    rows.append(row_values)
                text_parts.append(" ".join(row_values))

            sheet_data["sample_rows"] = rows[:5]
            sheet_data["row_count"] = sheet.max_row - 1 if sheet.max_row else 0
            result["sheets"][sheet_name] = sheet_data

        result["full_text"] = "\n".join(text_parts)
        wb.close()

    except Exception as e:
        result["error"] = str(e)

    return result


def analyze_docx(path: Path) -> Dict[str, Any]:
    """Analyze Word document."""
    result = {
        "paragraph_count": 0,
        "text_sample": "",
        "full_text": ""
    }

    if not HAS_DOCX:
        result["error"] = "python-docx not installed"
        return result

    try:
        doc = docx.Document(path)
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        result["paragraph_count"] = len(paragraphs)
        result["full_text"] = "\n".join(paragraphs[:50])
        result["text_sample"] = result["full_text"][:1000]

    except Exception as e:
        result["error"] = str(e)

    return result


def generate_content_summary(file_type: str, analysis: Dict[str, Any], entities: List[str], dates: List[str]) -> str:
    """Generate a human-readable content summary."""
    parts = []

    if file_type == "csv":
        headers = analysis.get("headers", [])
        row_count = analysis.get("row_count", 0)
        if headers:
            parts.append(f"{row_count} rows with columns: {', '.join(headers[:5])}")
            if len(headers) > 5:
                parts.append(f"(and {len(headers) - 5} more columns)")

    elif file_type == "pdf":
        page_count = analysis.get("page_count", 0)
        parts.append(f"{page_count} page PDF")
        text_sample = analysis.get("text_sample", "")[:200]
        if text_sample:
            # Clean up whitespace
            text_sample = " ".join(text_sample.split())
            parts.append(f"Content preview: {text_sample}...")

    elif file_type == "xlsx":
        sheets = analysis.get("sheet_names", [])
        parts.append(f"Excel workbook with {len(sheets)} sheet(s): {', '.join(sheets[:3])}")
        for sheet_name, sheet_data in analysis.get("sheets", {}).items():
            headers = sheet_data.get("headers", [])
            if headers:
                parts.append(f"'{sheet_name}' columns: {', '.join(headers[:5])}")

    elif file_type == "docx":
        para_count = analysis.get("paragraph_count", 0)
        parts.append(f"Word document with {para_count} paragraphs")
        text_sample = analysis.get("text_sample", "")[:200]
        if text_sample:
            text_sample = " ".join(text_sample.split())
            parts.append(f"Content preview: {text_sample}...")

    elif file_type == "text":
        line_count = analysis.get("line_count", 0)
        parts.append(f"Text file with {line_count} lines")
        text_sample = analysis.get("text_sample", "")[:200]
        if text_sample:
            text_sample = " ".join(text_sample.split())
            parts.append(f"Content preview: {text_sample}...")

    # Add entity info
    if entities:
        parts.append(f"Contains references to: {', '.join(entities)}")

    # Add date info
    if dates:
        parts.append(f"Dates found: {', '.join(dates[:3])}")

    return ". ".join(parts) if parts else "Unable to extract content summary"


def suggest_category(entities: List[str], file_type: str, analysis: Dict[str, Any]) -> Optional[str]:
    """Suggest filing category based on detected entities."""
    # Map entities to likely categories
    category_map = {
        "TD": "Financial - TD",
        "TD Visa": "Financial - TD Visa",
        "RBC": "Financial - RBC",
        "Tangerine": "Financial - Tangerine",
        "Wealthsimple": "Financial - Wealthsimple",
        "Rogers": "Utilities - Rogers",
        "Bell": "Utilities - Bell",
        "Fido": "Utilities - Fido",
        "Toronto Hydro": "Utilities - Hydro",
        "Hydro One": "Utilities - Hydro",
        "Enbridge": "Utilities - Gas",
        "TD Insurance": "Insurance - TD",
        "Intact": "Insurance",
        "Uken": "Work - Uken",
        "Cottage": "Real Estate - Cottage",
        "40 Gibson": "Real Estate - 40 Gibson",
        "Carlaw": "Real Estate - Carlaw",
    }

    # Return first matching category
    for entity in entities:
        if entity in category_map:
            return category_map[entity]

    return None


def analyze_file(file_path: str, max_lines: int = 50) -> Dict[str, Any]:
    """Main analysis dispatcher."""
    path = Path(file_path)

    if not path.exists():
        return {
            "error": f"File not found: {file_path}",
            "file_path": file_path,
        }

    # Get file info
    file_size = path.stat().st_size
    extension = path.suffix.lower()

    # Determine file type and analyze
    file_type_map = {
        '.csv': 'csv',
        '.tsv': 'csv',
        '.pdf': 'pdf',
        '.txt': 'text',
        '.md': 'text',
        '.log': 'text',
        '.json': 'text',
        '.xml': 'text',
        '.html': 'text',
        '.xlsx': 'xlsx',
        '.xls': 'xlsx',
        '.docx': 'docx',
        '.doc': 'docx',
    }

    file_type = file_type_map.get(extension, 'unknown')

    # Run appropriate analyzer
    if file_type == 'csv':
        analysis = analyze_csv(path, max_lines)
    elif file_type == 'pdf':
        analysis = analyze_pdf(path)
    elif file_type == 'text':
        analysis = analyze_text(path, max_lines)
    elif file_type == 'xlsx':
        analysis = analyze_xlsx(path)
    elif file_type == 'docx':
        analysis = analyze_docx(path)
    else:
        analysis = {"error": f"Unsupported file type: {extension}"}

    # Load entity config and detect entities
    entities_config = load_entity_config()
    full_text = analysis.get("full_text", "")
    detected_entities = detect_entities(full_text, entities_config) if full_text else []
    detected_dates = detect_dates(full_text) if full_text else []

    # Generate summary
    content_summary = generate_content_summary(file_type, analysis, detected_entities, detected_dates)

    # Suggest category
    suggested_category = suggest_category(detected_entities, file_type, analysis)

    # Calculate confidence boost based on what we found
    confidence_boost = 0
    if detected_entities:
        confidence_boost += 15  # Found known entities
    if detected_dates:
        confidence_boost += 5   # Found dates
    if full_text and len(full_text) > 100:
        confidence_boost += 5   # Successfully extracted substantial content

    # Build result (exclude full_text to keep output manageable)
    result = {
        "file_path": str(path.absolute()),
        "file_type": file_type,
        "file_size": file_size,
        "content_summary": content_summary,
        "detected_entities": detected_entities,
        "detected_dates": detected_dates,
        "suggested_category": suggested_category,
        "confidence_boost": confidence_boost,
        "analysis_details": {k: v for k, v in analysis.items() if k != "full_text"},
    }

    if "error" in analysis:
        result["error"] = analysis["error"]

    return result


def main():
    parser = argparse.ArgumentParser(
        description="Analyze file content for organization decisions"
    )
    parser.add_argument("file_path", help="Path to the file to analyze")
    parser.add_argument(
        "--max-lines",
        type=int,
        default=50,
        help="Maximum lines to read for text files (default: 50)"
    )
    parser.add_argument(
        "--pretty",
        action="store_true",
        help="Pretty-print JSON output"
    )

    args = parser.parse_args()

    result = analyze_file(args.file_path, args.max_lines)

    if args.pretty:
        print(json.dumps(result, indent=2))
    else:
        print(json.dumps(result))


if __name__ == "__main__":
    main()
